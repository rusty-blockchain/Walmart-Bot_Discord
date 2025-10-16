import pandas as pd
import os
import aiohttp
import csv
import json
import time
import concurrent.futures
from discord.ext import commands, tasks
from geopy.distance import geodesic
from dotenv import load_dotenv
import discord
import sqlite3
from datetime import datetime, timedelta, timezone
import pytz
from discord import app_commands
import discord
from discord.ext import commands
from discord import ui, Interaction
import re
from datetime import datetime, timedelta
import pytz
from discord.ext import tasks
from concurrent.futures import ThreadPoolExecutor
import asyncio
from pandas import ExcelWriter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import requests
from io import BytesIO
from PIL import Image as PILImage
from openpyxl.drawing.image import Image
import requests
from io import BytesIO
from concurrent.futures import ThreadPoolExecutor
import os
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.styles import Border, Side, Alignment
from openpyxl.worksheet.page import PageMargins
import colorsys
from openpyxl.styles import Font
from datetime import datetime, timedelta
from fpdf import FPDF
from fpdf.enums import XPos, YPos  # Add these imports
import random
image_executor = ThreadPoolExecutor(max_workers=2)


# After other imports
processing_executor = ThreadPoolExecutor(max_workers=2)

# Load environment variables from .env file
load_dotenv()

# Define intents
intents = discord.Intents.default()
intents.messages = True  # Enable the message event
intents.message_content = True  # Enable message content intent
intents.members = True  # Enable the members intent
intents.presences = True  # Enable the presence intent

# Path to the directory where CSV files will be stored
script_location = os.path.dirname(os.path.abspath(__file__))

temp_image_directory = os.path.join(script_location, 'temp_images')
if not os.path.exists(temp_image_directory):
    os.makedirs(temp_image_directory)
csv_directory = os.path.join(script_location, os.getenv('CSV_DIRECTORY'))
downloads_directory = os.path.join(csv_directory, 'downloads')
combined_csv_path = os.path.join(csv_directory, 'combined.csv')
combined_csv_path_temp = os.path.join(csv_directory, 'combinedtempfile.csv')
combined_csv_path_temp2 = os.path.join(csv_directory, 'combinednosalesid.csv')
zip_codes_csv_path = os.path.join(csv_directory, os.getenv('ZIP_CODES_CSV'))
zip_code_coordinates_csv_path = os.path.join(csv_directory, os.getenv('ZIP_CODE_COORDINATES_CSV'))
store_data_json_path = os.path.join(csv_directory, os.getenv('STORE_DATA_JSON'))
deals_channel_id = int(os.getenv('DEALS_CHANNEL_ID'))  # Replace with your deals channel ID
DEFAULT_IMAGE_URL = "https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcSo6HInE138_03esC-wW8ZXwgIh5Wc4VpAFQQ&s"
allowed_roles_setzip = os.getenv('ALLOWED_ROLES_SETZIP').split(',')
# Global dictionary to store UPC-item name mappings
upc_item_mappings = {}
failed_notifications = []
successful_notifications = 0
total_notifications = 0

# List to keep track of users the bot has DM'd
dm_users = []

# Ensure the downloads directory exists
if not os.path.exists(downloads_directory):
    os.makedirs(downloads_directory)
    


# Initialize the bot
bot = commands.Bot(command_prefix='/', intents=intents)
# Combine CSV files when the bot starts
@bot.event
async def on_ready():
    print(bot.user.id)
    try:
        await bot.tree.sync()
        print("Commands synchronized successfully.")
    except Exception as sync_e:
        print(f"Error syncing commands: {sync_e}")
    

    # Initialize the database
    conn = sqlite3.connect('user_data.db')
    cursor = conn.cursor()

    # Create the users table if it doesn't exist
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS users (
        username TEXT PRIMARY KEY,
        zip_code TEXT
    )
    ''')

    # Create the store_users table with the last_set_time column if it doesn't exist
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS store_users (
        username TEXT PRIMARY KEY,
        store_id TEXT,
        last_set_time TIMESTAMP DEFAULT '1970-01-01 00:00:00'
    )
    ''')
    
    # Check if the last_set_time column exists, and add it if it doesn't
    cursor.execute("PRAGMA table_info(store_users)")
    columns = [column[1] for column in cursor.fetchall()]
    if 'last_set_time' not in columns:
        cursor.execute('''
        ALTER TABLE store_users ADD COLUMN last_set_time TIMESTAMP DEFAULT '1970-01-01 00:00:00'
        ''')

    # Update the last_set_time for all existing users to the beginning of time
    cursor.execute('''
    UPDATE store_users SET last_set_time = '1970-01-01 00:00:00'
    ''')

    conn.commit()
    conn.close()
    
    await initialize_bot()
    # await notify_users_of_new_deals()
    
# Function to strip #0 from usernames in the database
def strip_username_suffix():
    conn = sqlite3.connect('user_data.db')
    cursor = conn.cursor()

    # Update usernames in the users table
    cursor.execute('''
        SELECT username FROM users WHERE username LIKE '%#0'
    ''')
    users_with_suffix = cursor.fetchall()
    for (username,) in users_with_suffix:
        new_username = username.replace('#0', '')
        cursor.execute('''
            SELECT COUNT(*) FROM users WHERE username = ?
        ''', (new_username,))
        count = cursor.fetchone()[0]
        if count == 0:
            cursor.execute('''
                UPDATE users
                SET username = ?
                WHERE username = ?
            ''', (new_username, username))
        else:
            print(f"Duplicate username found: {new_username}")

    # Update usernames in the store_users table
    cursor.execute('''
        SELECT username FROM store_users WHERE username LIKE '%#0'
    ''')
    store_users_with_suffix = cursor.fetchall()
    for (username,) in store_users_with_suffix:
        new_username = username.replace('#0', '')
        cursor.execute('''
            SELECT COUNT(*) FROM store_users WHERE username = ?
        ''', (new_username,))
        count = cursor.fetchone()[0]
        if count == 0:
            cursor.execute('''
                UPDATE store_users
                SET username = ?
                WHERE username = ?
            ''', (new_username, username))
        else:
            print(f"Duplicate username found: {new_username}")

    conn.commit()
    conn.close()
    
def remove_invalid_store_ids():
    conn = sqlite3.connect('user_data.db')
    cursor = conn.cursor()
    cursor.execute('SELECT username, store_id FROM store_users')
    store_users = cursor.fetchall()
    
    for username, store_id in store_users:
        if len(store_id) > 4:
            cursor.execute('DELETE FROM store_users WHERE username = ?', (username,))
    
    conn.commit()
    conn.close()

async def initialize_bot():
    if not os.path.exists(csv_directory):
        os.makedirs(csv_directory)
    
    clean_downloads_directory()
    deals_channel = bot.get_channel(deals_channel_id)
    print(f'Looking at Deals Channel')
    async for message in deals_channel.history(limit=100):
        if message.attachments and any(att.filename.endswith('.csv') for att in message.attachments):
            print(f'Found a CSV File')
            await fetch_csv_files_from_history(deals_channel)
            
            # Run CSV processing in separate thread
            success = await process_csv_operations_async()
            if success:
                print("CSV processing completed successfully")
            break
    strip_username_suffix()
    await send_zip_code_prompt()
    await send_store_id_prompt()
    
    send_prompts.start()
    remove_invalid_store_ids()
    auto_rebuild.start()
    daily_deals_notification.start()
    print(f'Bot is ready and logged in as {bot.user}')


import csv
import os

import csv
import os

# Load the store data from the JSON file
with open('csv/STORE_DATA_JSON.json', 'r') as json_file:
    store_data = json.load(json_file)

# Create a dictionary to map postal codes to store ids
postal_code_to_store_id = {store['postal_code']: store['store_id'] for store in store_data}

async def process_csv_operations_async():
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(
        processing_executor,
        process_csv_operations_sync,
        downloads_directory,
        combined_csv_path_temp,
        combined_csv_path_temp2,
        combined_csv_path
    )


async def report_failed_notifications(channel_id):
    channel = bot.get_channel(channel_id)    
    if not failed_notifications:
        if channel:
            embed = discord.Embed(
                title="Daily Notifications Sent",
                description=f"All Daily Notifications Sent Successfully\nTotal Successful Notifications: {successful_notifications} Out of {total_notifications}",
                color=discord.Color.green()
            )
            await channel.send(embed=embed)
    
    if channel:
        failed_users = "\n".join([f"- {user}" for user in failed_notifications])
        embed = discord.Embed(
            title="Failed Notifications Report",
            description=f"The following users could not be notified:\n{failed_users}\nTotal Successful Notifications: {successful_notifications} Out of {total_notifications}",
            color=discord.Color.red()
        )
        view = RetryNotificationButton(failed_notifications.copy())
        await channel.send(embed=embed, view=view)
        failed_notifications.clear()


def process_csv_operations_sync(downloads_dir, temp_path, temp2_path, final_path):
    try:
        # Run operations sequentially
        process_and_combine_csv(downloads_dir, temp_path)
        reorganize_columns(temp_path, temp2_path)
        update_store_ids(temp2_path, final_path)
        return True
    except Exception as e:
        print(f"Error in CSV processing: {e}")
        return False

# Function to remove user from the database
def remove_user_from_db(username):
    conn = sqlite3.connect('user_data.db')
    cursor = conn.cursor()
    cursor.execute('DELETE FROM users WHERE username = ?', (username,))
    cursor.execute('DELETE FROM store_users WHERE username = ?', (username,))
    conn.commit()
    conn.close()

def generate_light_colors(n):
    colors = []
    for i in range(n):
        hue = i / n
        # High saturation (0.3) and value (0.95) for light but visible colors
        rgb = colorsys.hsv_to_rgb(hue, 0.3, 0.95)
        # Convert to hex color
        hex_color = '%02x%02x%02x' % (int(rgb[0]*255), int(rgb[1]*255), int(rgb[2]*255))
        colors.append(hex_color)
    return colors

def calculate_discount(msrp, price):
    try:
        # Strip whitespace and handle empty/N/A values
        if not msrp or msrp.strip() in ['', 'N/A']:
            return None
        if not price or price.strip() in ['', 'N/A']:
            return None
            
        # Convert to float and calculate
        msrp = float(msrp.strip())
        price = float(price.strip())
        
        if msrp == 0:
            return 100
        else:
            discount = ((msrp - price) / msrp) * 100
            return round(discount)
    except (ValueError, TypeError, AttributeError):
        return None
        
        
async def extract_sku_from_embed(channel, csv_message):
    async for message in channel.history(limit=100, before=csv_message):
        if message.embeds:
            for embed in message.embeds:
                # Check if the embed has the title "Stock Information Walmart"
                if embed.title == "Stock Information Walmart":
                    return None, None  # Stop searching when the titled embed is found
                # Look for the SKU in the embed fields
                sku = None
                image_url = None
                for field in embed.fields:
                    if field.name == "SKU (UPC)":
                        sku = field.value.strip()
                        # Remove parentheses and everything inside them
                        sku = re.sub(r'\(.*?\)', '', sku)
                        # Add the URL to the beginning of the SKU
                        sku = f"https://www.walmart.com/ip/emoney/{sku}"
                # Get the image URL if available
                if embed.thumbnail and embed.thumbnail.url:
                    image_url = embed.thumbnail.url
                #if sku and image_url:
                    return None, image_url
    return None, None

async def get_upc_from_embed(channel, csv_message):
    """
    Extracts UPC and SKU from Discord embed messages.
    Returns: Tuple of (UPC string, SKU string) if found, (None, None) if not found
    """
    async for message in channel.history(limit=100, before=csv_message):
        if message.embeds:
            for embed in message.embeds:
                for field in embed.fields:
                    if field.name == "SKU (UPC)":
                        full_value = field.value.strip()
                        upc_match = re.search(r'\((.*?)\)', full_value)
                        if upc_match:
                            sku = re.sub(r'\(.*?\)', '', full_value).strip()
                            sku = f"https://www.walmart.com/ip/emoney/{sku}"
                            return upc_match.group(1), sku
    return None, None

        

def process_and_combine_csv(input_folder, output_file):
    # Create directories if they don't exist
    os.makedirs(input_folder, exist_ok=True)
    os.makedirs(os.path.dirname(output_file), exist_ok=True)    
    for file_name in os.listdir(input_folder):
        if file_name.endswith('.csv'):
            file_path = os.path.join(input_folder, file_name)
    global upc_item_mappings
    def process_csv(file_name):
        print(f"Processing file: {file_name}")
        
        with open(file_name, 'r', encoding='utf-8') as file:
            reader = csv.reader(file, quotechar='"', delimiter=',', quoting=csv.QUOTE_MINIMAL)
            rows = list(reader)
            
            
            
                # Initialize variables
            url = ""
            item_name = ""
            msrp = "0.00"
            image_url = ""
            
        
        
                # Ensure 'Aisles' column is present
        if ' Aisles' not in rows[0]:
            rows[0].append(' Aisles')
        for row in rows[1:]:
            row.append('')
        
            header = rows[0]
# Add leading zeros to ZIP codes if they have less than 5 digits
        if ' ZIP' in header:
            zip_index = header.index(' ZIP')
            for i, row in enumerate(rows[1:]):
                if i == len(rows) - 2:
                    continue  # Skip the last row when processing ZIP codes
                zip_code = ''.join(filter(str.isdigit, row[zip_index]))  # Remove non-numeric values
                rows[i + 1][zip_index] = f"{int(zip_code):05}" if zip_code else ''
        
        
        # Combine data in empty header columns into the 'Aisles' column
        header = rows[0]

        aisle_index = header.index(' Aisles')
        for row in rows[1:]:
            combined_data = []
            for i, value in enumerate(row):
                if i > aisle_index and i < len(row) and value:
                    combined_data.append(value.strip())
                    row[i] = ''
            if combined_data:
                row[aisle_index] += '/' + '/'.join(combined_data)
            # Ensure the index exists before accessing it
            if aisle_index < len(row):
                if row[aisle_index].strip() == 'N/A':
                    row[aisle_index] = 'Possibly OOS'
                row[aisle_index] = row[aisle_index].rstrip(',')
            
            
            
        # Check if the last line contains product information
        last_line = rows[-1]
        last_line_str = ','.join(last_line)
        if 'MSRP' in last_line_str:
            url_parts = last_line_str.split(',')
            # Find the part containing "MSRP:"
            msrp_part = None
            for part in url_parts:
                if 'MSRP:' in part:
                    msrp_part = part
                    msrp_index = url_parts.index(part)
                    break
                    
            if msrp_part:
                upc = url_parts[0].strip()
                url = url_parts[1].strip()
                image_url = url_parts[2].strip()
                item_name = url_parts[3].strip()
                msrp = msrp_part.split('MSRP:')[1].strip()
                
                
                
                
                # Debug prints
                print(f"DEBUG: Processing last line")
                print(f"UPC: {upc}")
                print(f"URL: {url}")
                print(f"Image URL: {image_url}")
                print(f"Item Name: {item_name}")
                print(f"MSRP: {msrp}")
                upc_item_mappings[item_name] = upc
                rows.pop()  # Remove the last line
            else:
                print("Last line does not contain enough parts.")
        else:
            print("MSRP not found in the last line.")
        
        # Add item_name, MSRP, and URL to the beginning of each row
        for row in rows[1:]:
            row.insert(0, image_url)
            row.insert(0, url)
            row.insert(0, float(msrp))
            row.insert(0, item_name)
            
        
        # Ensure 'item_name', 'MSRP', and 'URL' columns are present
        if 'item_name' not in rows[0]:
            rows[0].insert(0, 'item_name')
        if 'MSRP' not in rows[0]:
            rows[0].insert(1, 'MSRP')
        if 'URL' not in rows[0]:
            rows[0].insert(2, 'URL')
        if 'image_url' not in rows[0]:
            rows[0].insert(3, 'image_url')
        
        # Ensure 'Store ID' column is present and filled correctly
        if 'Store ID' not in rows[0]:
            rows[0].insert(4, 'Store ID')
        for row in rows[1:]:
            row.insert(4, '0')
        
        
        
            
        for i, row in enumerate(rows):
            rows[i] = [str(value).replace(',', '') for value in row]
            rows[i] = [str(value).replace("'", '') for value in row]
            
        # Calculate discount and add the column after 'Store ID'
        if 'MSRP' in rows[0] and ' Price' in rows[0]:
            msrp_index = rows[0].index('MSRP')
            price_index = rows[0].index(' Price')
            rows[0].append('discount')
            discount_index = rows[0].index('discount')
            
            
            
                    # Process rows
        for row in rows[1:]:  # Skip header
            if len(row) >= 2:  # Make sure row has enough columns
                msrp = row[msrp_index] if msrp_index >= 0 and msrp_index < len(row) else '0'
                price = row[price_index] if price_index >= 0 and price_index < len(row) else '0'
                
                # Calculate discount
                
                discount = calculate_discount(msrp, price)
                
                # Add discount to row, use '0%' if None
                
                row[discount_index] = str(discount) + '%' if discount is not None else '0%'
            else:
                print(f"Row {row} does not have a 'Price' column.")
                row.insert(discount_index, '0%')
        
        
        
        # Remove anything after Aisles
        # del row[aisle_index + 1:]
        # row[aisle_index] = row[aisle_index].rstrip(',')
         # When adding UPC column, use the mapping
        #if 'upc' not in rows[0]:
           # rows[0].insert(0, 'upc')
           # for row in rows[1:]:
              #  item_name = row[rows[0].index('item_name')]
              #  row.insert(0, upc_item_mappings.get(item_name, ''))


        #Strip commas from non-address columns
        for i, row in enumerate(rows):
            rows[i] = [str(value).replace(',', '') for value in row]
        
        # Move columns to the left for rows with 'PR' in the ZIP code column
        for row in rows[1:]:
            if 'PR' in row[header.index(' ZIP')]:
                row[header.index(' ZIP')] = row[header.index(' Backroom Stock')]
                row[header.index(' Backroom Stock')] = row[header.index(' Floor Stock')]
                row[header.index(' Floor Stock')] = row[header.index(' In Transit Stock')]
                row[header.index(' In Transit Stock')] = row[header.index(' Price')]
                row[header.index(' Price')] = row[header.index(' Aisles')]
                row.pop()  # Remove the last column
         
        return rows
    
    seven_days_ago = datetime.now() - timedelta(days=2)  
    
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    combined_data = []
    for file_name in os.listdir(input_folder):
        file_path = os.path.join(input_folder, file_name)
        if file_name.endswith('.csv') and file_name != "combined_temp.csv":
            
            
                
            processed_rows = process_csv(os.path.join(input_folder, file_name))
            if not combined_data:
                combined_data.append(processed_rows[0])
            combined_data.extend(processed_rows[1:])
    
    with open(output_file, 'w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file, quotechar='"', delimiter=',', quoting=csv.QUOTE_MINIMAL)
        writer.writerows(combined_data)
    
    print(f"All files have been combined and saved as '{output_file}'.")



def reorganize_columns(input_csv, output_csv):
    with open(input_csv, 'r', encoding='utf-8') as infile, open(output_csv, 'w', newline='', encoding='utf-8') as outfile:
        reader = csv.DictReader(infile)
        fieldnames = ['image_url', 'Store ID', 'item_name', 'MSRP', ' Price', 'discount', 
                      ' Floor Stock', ' Backroom Stock', ' In Transit Stock', 
                      ' Aisles', 'URL', 'Address', ' City', ' State', ' ZIP']
        
        writer = csv.DictWriter(outfile, fieldnames=fieldnames)
        writer.writeheader()
        
        for row in reader:
            writer.writerow({field: row.get(field, '') for field in fieldnames})
    
    print(f"Columns have been reorganized and saved as '{output_csv}'.")
    
def update_store_ids(input_csv, output_csv):
    with open(input_csv, 'r', encoding='utf-8') as infile, open(output_csv, 'w', newline='', encoding='utf-8') as outfile:
        reader = csv.reader(infile)
        writer = csv.writer(outfile)
        # Read the header
        try:
            header = next(reader)
            writer.writerow(header)

            # Find the index of the Store ID and ZIP columns
            store_id_index = header.index('Store ID')
            zip_index = header.index(' ZIP')

            # Process each row
            for row in reader:
                if row[store_id_index] == '0':
                    postal_code = row[zip_index].strip()
                    #print(f"Checking postal code: {postal_code}")
                    if postal_code in postal_code_to_store_id:
                        row[store_id_index] = postal_code_to_store_id[postal_code]
                        #print(f"Updated Store ID for postal code {postal_code}: {row[store_id_index]}")
                writer.writerow(row)
        except StopIteration:
            print("The input CSV file is empty or does not contain any data.")
    
async def fetch_user_by_name(username):
    for guild in bot.guilds:
        for member in guild.members:
            if member.name == username:
                return member
    return None

@bot.command(name='list_dm_users')
async def list_dm_users(ctx):
    if dm_users:
        await user.send("Users the bot has DM'd:\n" + "\n".join(dm_users))
    else:
        await user.send("The bot has not DM'd any users yet.")

required_role_name = os.getenv('ALLOWED_ROLES_SETZIP').split(',')  # Replace with your required role name


#function to notify users of new deals
async def notify_users_of_new_deals():
    print("Starting notify_users_of_new_deals function")
    conn = sqlite3.connect('user_data.db')
    cursor = conn.cursor()
    
    print("Loading combined CSV file")
    try:
        combined_df = pd.read_csv(combined_csv_path, dtype={' ZIP': str, 'Store ID': str})
        print(f"Successfully loaded CSV with {len(combined_df)} rows")
        print("Sample of data:")
        print(combined_df.head())
    except Exception as e:
        print(f"Error loading combined CSV file: {e}")
        return

    print("Converting discount column")
    def convert_discount(x):
        try:
            return int(x.rstrip('%'))
        except (ValueError, AttributeError):
            return 0
    combined_df['discount'] = combined_df['discount'].apply(convert_discount)

    print("Converting stock columns to numeric")
    stock_columns = [' Floor Stock', ' Backroom Stock', ' In Transit Stock']
    for col in stock_columns:
        combined_df[col] = pd.to_numeric(combined_df[col], errors='coerce').fillna(0).astype(int)

    print("Fetching users with zip codes")
    cursor.execute('SELECT username, zip_code FROM users')
    user_zip_codes = cursor.fetchall()
    print(f"Found {len(user_zip_codes)} users with zip codes")

    # Add randomization to avoid all DMs being sent at exactly the same time
    user_zip_codes = list(user_zip_codes)
    random.shuffle(user_zip_codes)

    for idx, (username, zip_code) in enumerate(user_zip_codes):
        print(f"\nProcessing user {idx + 1}/{len(user_zip_codes)}: {username} with zip code: {zip_code}")
        user = await fetch_user_by_name(username)
        
        if user:
            print(f"User object found for {username}")
            # Check if user has required role
            user_roles = [str(role.id) for role in user.roles]
            if not any(role_id in allowed_roles_setzipbutton for role_id in user_roles):
                print(f"User {username} no longer has required role. Skipping.")
                # Optionally remove user from database if they don't have the role anymore
                
                
            
            print(f"User roles: {user_roles}")
            
            print(f"Finding nearby zip codes within 50 miles")
            nearby_zip_codes = storecheck(zip_code, 50)
            print(f"Found {len(nearby_zip_codes)} nearby zip codes")

            print("Filtering deals")
            filtered_df = combined_df[
                (combined_df[' ZIP'].str[:5].isin([z[:5] for z in nearby_zip_codes])) &
                (combined_df['discount'] >= 50) &
                (
                    (combined_df[' Floor Stock'] > 1) |
                    (combined_df[' Backroom Stock'] > 1) |
                    (combined_df[' In Transit Stock'] > 1)
                )
            ]
            print(f"Found {len(filtered_df)} matching deals")

            if not filtered_df.empty:
                print(f"Processing deals for {username}")
                output_file_path = os.path.join(user_csv_directory, f'user_{username}_deals.csv')
                filtered_df.to_csv(output_file_path, index=False)
                print(f"Saved filtered deals to {output_file_path}")
                
                print(f"Sending deals to {username}")
                max_retries = 3
                retry_delay = 5  # seconds
                
                for attempt in range(max_retries):
                    try:
                        await add_percentage_and_send(output_file_path, user)
                        print(f"Successfully sent deals to {username}")
                        break
                    except discord.errors.HTTPException as e:
                        if e.status == 429:  # Rate limit error
                            retry_after = e.retry_after if hasattr(e, 'retry_after') else retry_delay
                            print(f"Rate limited when sending to {username}. Waiting {retry_after} seconds...")
                            await asyncio.sleep(retry_after)
                            continue
                        else:
                            print(f"HTTP error when sending to {username}: {e}")
                            failed_notifications.append(username)
                            break
                    except Exception as e:
                        print(f"Error sending to {username}: {e}")
                        failed_notifications.append(username)
                        break
                else:
                    print(f"Failed to send to {username} after {max_retries} attempts")
                    failed_notifications.append(username)
            else:
                print(f"No deals found for {username}")
                try:
                    await user.send(f"No deals found within your radius.")
                except discord.errors.Forbidden:
                    print(f"Cannot send message to {username} - DMs might be disabled")
                    failed_notifications.append(username)
                except Exception as e:
                    print(f"Error sending no-deals message to {username}: {e}")
                    failed_notifications.append(username)
        else:
            print(f"Could not find user object for {username}")
            failed_notifications.append(username)
            # Remove user from database if they can't be found
            cursor.execute('DELETE FROM users WHERE username = ?', (username,))
            conn.commit()
            
        # Add a random delay between users to avoid rate limits
        delay = random.uniform(1.5, 3.0)  # Random delay between 1.5 and 3 seconds
        print(f"Waiting {delay:.2f} seconds before processing next user...")
        await asyncio.sleep(delay)

    notification_channel_id = int(os.getenv('NOTIFICATION_CHANNEL_ID'))
    await report_failed_notifications(notification_channel_id)
    print("Notification process complete")
    conn.close()

async def download_file(url, dest, channel, csv_message):
    async with aiohttp.ClientSession() as session:
        async with session.get(url) as response:
            with open(dest, 'wb') as f:
                f.write(await response.read())

    # Extract SKU and image URL from the closest embed message
    unused, image_url = await extract_sku_from_embed(channel, csv_message)
    upc, sku = await get_upc_from_embed(channel, csv_message)
    

    # Process the CSV file to add the SKU and image URL to the beginning of the last row
    with open(dest, 'r', encoding='utf-8') as file:
        reader = csv.reader(file, quotechar='"', delimiter=',', quoting=csv.QUOTE_MINIMAL)
        rows = list(reader)

    if rows:

        
        last_row = rows[-1]
        last_row.insert(0, image_url)
        last_row.insert(0, sku)
        last_row.insert(0, upc)
        print(last_row)
        


    with open(dest, 'w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file, quotechar='"', delimiter=',', quoting=csv.QUOTE_MINIMAL)
        writer.writerows(rows)

def clean_downloads_directory():
    import time
    
    print("Starting directory cleanup...")
    if os.path.exists(downloads_directory):
        for file in os.listdir(downloads_directory):
            file_path = os.path.join(downloads_directory, file)
            max_attempts = 3
            attempt = 0
            
            while attempt < max_attempts:
                try:
                    os.remove(file_path)
                    print(f"Deleted: {file}")
                    break
                except PermissionError:
                    print(f"File in use, waiting: {file}")
                    time.sleep(1)
                    attempt += 1
                except Exception as e:
                    print(f"Error deleting {file}: {e}")
                    break
    print("Cleanup complete")

async def fetch_csv_files_from_history(channel):
    
    # Calculate timestamp for 7 days ago
    seven_days_ago = datetime.now(timezone.utc) - timedelta(days=1)
    print(f"Fetching messages after: {seven_days_ago}")
    
    async for message in channel.history(limit=None):
        # Stop if message is older than 7 days
        if message.created_at < seven_days_ago:
            print(f"Reached message older than 7 days, stopping fetch")
            break
            
        if message.attachments:
            for attachment in message.attachments:
                if attachment.filename.endswith('.csv'):
                    print(f"Found CSV from {message.created_at}: {attachment.filename}")
                    file_path = os.path.join(downloads_directory, attachment.filename)
                    await download_file(attachment.url, file_path, channel, message)

    print("Finished fetching CSV files")

@bot.event
async def on_message(message):
    channel = bot.get_channel(deals_channel_id)
    if message.channel.id == deals_channel_id and message.attachments:
        for attachment in message.attachments:
            if attachment.filename.endswith('.csv'):
                file_path = os.path.join(downloads_directory, attachment.filename)
                await download_file(attachment.url, file_path, channel, message)
                
                # Extract SKU from the closest embed message
                sku = await extract_sku_from_embed(message.channel, message)
                
                # Process and combine CSV files
                combined_csv_path_temp = os.path.join(csv_directory, 'combined_temp.csv')
                process_and_combine_csv(downloads_directory, combined_csv_path_temp)
                reorganize_columns(combined_csv_path_temp, combined_csv_path_temp2)
                update_store_ids(combined_csv_path_temp2, combined_csv_path)
                
                await bot.process_commands(message)

# Your storecheck function
def storecheck(zip_code, radius):
    results = []
    def load_zip_code_coordinates(zip_codes_csv):
        zip_code_coords = {}
        with open(zip_codes_csv, 'r') as csvfile:
            print(f"ZIP CODE CSV FILE OPEN")
            csvreader = csv.reader(csvfile)
            next(csvreader)  # Skip the header row
            for row in csvreader:
                zip_code, lat, lon = row[0], float(row[1]), float(row[2])
                zip_code_coords[zip_code] = (lat, lon)
                #print(zip_code_coords)
        return zip_code_coords

    def find_zip_codes_within_radius(zip_code, radius, zip_code_coords):
        #print(zip_code)
        origin_coords = zip_code_coords.get(zip_code)
        #print(origin_coords)
        if not origin_coords:
            return []
        nearby_zip_codes = []
        for other_zip_code, coords in zip_code_coords.items():
            distance = geodesic(origin_coords, coords).miles
            if distance <= radius and other_zip_code != zip_code:
                nearby_zip_codes.append(other_zip_code)
        print(nearby_zip_codes)
        return nearby_zip_codes

    zip_code_coords = load_zip_code_coordinates(zip_code_coordinates_csv_path)
    target_zip_code = zip_code
    if radius > 50:
        radius = 50  # 50 miles
    nearby_zip_codes = find_zip_codes_within_radius(target_zip_code, radius, zip_code_coords)
    return nearby_zip_codes

# Define the new directory for user-specific CSV files
user_csv_directory = os.path.join(script_location, 'user_csv_files')
# Ensure the user_csv_directory exists
if not os.path.exists(user_csv_directory):
    os.makedirs(user_csv_directory)

# Function to delete a file
def delete_file(file_path):
    try:
        os.remove(file_path)
        print(f"Deleted file: {file_path}")
    except OSError as e:
        print(f"Error deleting file {file_path}: {e}")

def compress_image(pil_image, max_size=(50, 50)):
    # Resize image while maintaining aspect ratio
    pil_image.thumbnail(max_size, PILImage.Resampling.LANCZOS)
    # Compress image
    output = BytesIO()
    pil_image.save(output, format='PNG', optimize=True, quality=65)
    return output




def process_excel_with_images_sync(excel_path, image_urls):
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        }
        print("Starting Excel processing...")
        wb = load_workbook(excel_path)
        ws = wb.active
        print("Workbook loaded successfully")

        # Find eBay Link column
        ebay_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'Ebay Link':
                ebay_col = idx
                print(f"Found Ebay Link column at index {idx}")
                break

        # Update eBay links
        if ebay_col:
            print("Processing eBay links...")
            for row in ws.iter_rows(min_row=2):  # Skip header
                cell = row[ebay_col-1]  # -1 because column index is 0-based
                if cell.value and isinstance(cell.value, str) and 'ebay.com' in cell.value:
                    cell.hyperlink = cell.value
                    cell.font = Font(color="0000FF", underline="single")
                    cell.value = 'eBay Link'
            print("eBay links processed")

        # Find Walmart Link column (similar to how we found eBay Link)
        walmart_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'URL':  # Looking for the URL column
                walmart_col = idx
                print(f"Found Walmart Link column at index {idx}")
                break

        # Update Walmart links
        if walmart_col:
            print("Processing Walmart links...")
            for row in ws.iter_rows(min_row=2):  # Skip header
                cell = row[walmart_col-1]
                if cell.value and isinstance(cell.value, str) and 'walmart.com' in cell.value:
                    # Strip any whitespace and %20 from the URL
                    clean_url = cell.value.strip().rstrip('%20')
                    cell.hyperlink = clean_url
                    cell.font = Font(color="0000FF", underline="single")
                    cell.value = 'Walmart Link'
        print("Walmart links processed")        

        # Image processing
        print("Processing images...")
        image_url_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'image_url':
                image_url_col = idx
                print(f"Found image_url column at index {idx}")
                break
        
        for idx, url in enumerate(image_urls, start=2):
            try:
                response = requests.get(url, headers=headers)
                print(f"Image URL response status: {response.status_code}")
                if response.status_code != 200:
                    print(f"Using default image for index {idx-1}")
                    response = requests.get(DEFAULT_IMAGE_URL)
                    
                img_data = BytesIO(response.content)
                pil_image = PILImage.open(img_data)
                print(f"Successfully loaded image {idx-1}")
                
                output = compress_image(pil_image)
                img = Image(output)
                
                img.width = 40
                img.height = 40
                ws.row_dimensions[idx].height = 35
                ws.add_image(img, f'A{idx}')
                print(f"Added image {idx-1} to Excel")
                
                if image_url_col:
                    ws.cell(row=idx, column=image_url_col).value = ''
                    
            except Exception as e:
                print(f"Error processing image {idx-1}: {e}")
                continue

        print("Starting color coding...")
        df = pd.read_excel(excel_path)
        unique_stores = df['Store ID'].unique()
        print(f"Found {len(unique_stores)} unique store IDs")
        colors = generate_light_colors(len(unique_stores))
        store_colors = dict(zip(unique_stores, colors))
        
        for row in range(2, ws.max_row + 1):
            store_id = ws.cell(row=row, column=df.columns.get_loc('Store ID') + 1).value
            fill = PatternFill(start_color=store_colors[store_id], 
                             end_color=store_colors[store_id],
                             fill_type='solid')
            for cell in ws[row]:
                cell.fill = fill
        print("Finished color coding")


        print("Setting page orientation...")
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = False
        print("Page orientation set to landscape")
        ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25)
        print("Adding borders...")
        thin_border = Border(left=Side(style='thin'), 
                           right=Side(style='thin'), 
                           top=Side(style='thin'), 
                           bottom=Side(style='thin'))

        cell_count = 0
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='center')
                cell_count += 1
        print(f"Added borders to {cell_count} cells")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 0.9  # Reduce width by 10%
            ws.column_dimensions[column_letter].width = min(adjusted_width, 30)  # Cap width at 30
            
        # Adjust row heights
        for row in ws.rows:
            ws.row_dimensions[row[0].row].height = 30
            
        print("Cell formatting complete")

        print(f"Saving workbook to {excel_path}")
        wb.save(excel_path)
        print("Excel processing complete")

    except Exception as e:
        print(f"Error in process_excel_with_images_sync: {e}")
        print(f"Error occurred at step: {e.__traceback__.tb_lineno}")
        raise



def clean_dataframe(df):
    # Function to clean individual strings
    def clean_text(text):
        if isinstance(text, str):
            return ''.join(char for char in text if ord(char) < 256)
        return text
    
    # Apply cleaning to all string columns
    return df.applymap(clean_text)




def create_pdf_report(excel_path, pdf_path, df):
    pdf = FPDF(orientation='L', format='A4')
    pdf.add_page()
    # Use built-in fonts instead of custom fonts
    # pdf.add_font("dejavu-sans-narrow", style="", fname="DejaVuSansCondensed.ttf", uni=True)
    # pdf.add_font("dejavu-sans-narrow", style="b", fname="DejaVuSansCondensed-Bold.ttf", uni=True)
    pdf.set_auto_page_break(auto=True, margin=15)

    
    
    
    # Calculate optimal dimensions
    page_width = pdf.w - 30
    img_width = 25
    row_height = 25
    
    # Include all columns with optimized widths
    remaining_width = page_width - img_width
    
    # Dynamic width allocation based on content type
    col_widths = {
        'Store ID': 0.06,
        'item_name': 0.15,
        'MSRP': 0.05,
        ' Price': 0.05,
        'discount': 0.06,
        ' Floor Stock': 0.05,
        ' Backroom Stock': 0.05,
        ' In Transit Stock': 0.05,
        ' Aisles': 0.08,
        'URL': 0.05,
        'Ebay Link': 0.06,
        'UPC': 0.08,
        'Address': 0.10,
        ' City': 0.06,
        ' State': 0.04,
        ' ZIP': 0.06
    }
    
    # Generate colors for each unique store ID
    unique_stores = df['Store ID'].unique()
    colors = generate_light_colors(len(unique_stores))
    store_colors = dict(zip(unique_stores, colors))
    
    # Header styling with uniform height
    pdf.set_font("Arial", 'B', 8)
    pdf.set_fill_color(230, 230, 230)
    
    # Draw headers with consistent height
    max_header_height = row_height
    y_start = pdf.get_y()
    x_start = pdf.get_x()
    
    # Image header - single line
    pdf.cell(img_width, max_header_height, "Image", border=1, align='C', fill=True)
    pdf.ln(0)
    current_x = x_start + img_width
    
    # Column headers - all same height
    for col, width_ratio in col_widths.items():
        width = remaining_width * width_ratio
        pdf.set_xy(current_x, y_start)
        
        # Format header text with strategic line breaks
        header_text = col.strip()
        if 'Floor Stock' in header_text:
            header_text = 'Floor'
        elif 'Backroom Stock' in header_text:
            header_text = 'Backroom'
        elif 'In Transit Stock' in header_text:
            header_text = 'In Transit' 
        elif 'Ebay Link' in header_text:
            header_text = 'Ebay'

        pdf.cell(width, max_header_height, header_text, border=1, align='C', fill=True)
        current_x += width
        
    
    pdf.ln(max_header_height)
    
    # Data rows
    pdf.set_font("Arial", '', 7)
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    }

    for idx, row in df.iterrows():
        row_start_y = pdf.get_y()
        
        if row_start_y + row_height > pdf.page_break_trigger:
            pdf.add_page()
            row_start_y = pdf.get_y()
        
        try:
            item_name = row['item_name']
            upc_value = upc_item_mappings.get(item_name, 'No UPC Found')

            # Set background color based on store ID
            store_id = row['Store ID']
            color = store_colors[store_id]
            r = int(color[:2], 16)
            g = int(color[2:4], 16)
            b = int(color[4:], 16)
            pdf.set_fill_color(r, g, b)
            
            # Image cell
            x_start = pdf.get_x()
            image_url = row['image_url']
            if isinstance(image_url, (int, float)):
                    image_url = str(image_url)
            if image_url and image_url.strip():
                try:
                    response = requests.get(image_url, headers=headers, timeout=10)
                    if response.status_code == 200:
                        temp_img_path = os.path.join(temp_image_directory, f"temp_img_{idx}.png")
                        with open(temp_img_path, 'wb') as f:
                            f.write(response.content)
                        
                        with PILImage.open(temp_img_path) as img:
                            img = img.resize((100, 100), PILImage.LANCZOS)
                            img = img.convert('RGB')
                            # Get original image format
                            img_format = img.format.lower() if img.format else 'png'
                            
                            # Save in original format if supported, otherwise default to PNG
                            if img_format in ['png', 'jpeg', 'jpg']:
                                img.save(temp_img_path, format=img_format.upper())
                            else:
                                img.save(temp_img_path, format='PNG')
                        
                        pdf.image(temp_img_path, x=x_start + 2, y=row_start_y + 2,
                                w=img_width - 4, h=row_height - 4)
                        os.remove(temp_img_path)
                except Exception as e:
                    print(f"Image processing error for {image_url}: {e}")
            
            pdf.rect(x_start, row_start_y, img_width, row_height)
            pdf.set_xy(x_start + img_width, row_start_y)
            
            # Data cells with color fill
            for col, width_ratio in col_widths.items():
                width = remaining_width * width_ratio
                value = str(row.get(col, ''))

                x_pos = pdf.get_x()
                if col == 'UPC':
                    pdf.cell(width, row_height, upc_value, border=1, align='C', fill=True)
                elif col in ['URL', 'Ebay Link'] and value.startswith('http'):
                    pdf.set_text_color(0, 0, 255)
                    pdf.cell(width, row_height, 'Link', border=1, align='C', fill=True, link=value.strip())
                    pdf.set_text_color(0, 0, 0)
                else:
                    if len(value) > 20 and col not in ['UPC']:
                        value = value[:17] + '...'
                    pdf.cell(width, row_height, value, border=1, align='C', fill=True)
                pdf.set_xy(x_pos + width, row_start_y)
            
            pdf.ln(row_height)
            
        except Exception as e:
            print(f"Error processing row: {e}")
            continue
    
    pdf.output(pdf_path)


async def add_percentage_and_send(csv_file_path, user):
    global total_notifications
    print(f"Starting add_percentage_and_send for {user.name}")
    try:
        print("Reading CSV file")
        df = pd.read_csv(csv_file_path)
        df = df[df['discount'] != 100]
        if 'discount' in df.columns:
            print("Converting discount column")
            df['discount'] = df['discount'].astype(str) + '%'
            df['discount'] = df['discount'].str.rstrip('%').astype(float)

         # Sort and remove duplicates here
        df = df.sort_values(['Store ID', 'discount'], ascending=[True, False])
        df = df.drop_duplicates(subset=['Store ID', 'item_name'], keep='first')

         # Replace MSRP of 0 with 'N/A'
        df['MSRP'] = df['MSRP'].apply(lambda x: 'N/A' if float(x) == 0 else x)
        
        print("Creating eBay links")
        # Inside add_percentage_and_send:
        df['Ebay Link'] = df['item_name'].astype(str).apply(
            lambda x: f"https://www.ebay.com/sch/i.html?_nkw={x.replace(' ', '+')}&_sacat=0&rt=nc&LH_Sold=1&LH_Complete=1"
        )       
        
        print("Sorting data")
        df = df.sort_values(['Store ID', 'discount'], ascending=[True, False])
        df['discount'] = df['discount'].astype(str) + '%'
        
        excel_path = csv_file_path.replace('.csv', '.xlsx')
        print(f"Saving to Excel: {excel_path}")
        df.to_excel(excel_path, index=False)
        
        image_urls = df['image_url'].tolist()
        print(f"Processing {len(image_urls)} images")
        loop = asyncio.get_running_loop()
        await loop.run_in_executor(image_executor, 
                                 process_excel_with_images_sync, 
                                 excel_path, 
                                 image_urls)
        
        try:
            print("Converting to PDF : ")
            
            #df = pd.read_csv(csv_file_path)
            # Clean the DataFrame before PDF generation
            print("Before cleaning:")
            
            df = clean_dataframe(df)
            print("After cleaning:")
            pdf_path = csv_file_path.replace('.csv', '.pdf')
            df['Ebay Link'] = df['item_name'].astype(str).apply(
                lambda x: f"https://www.ebay.com/sch/i.html?_nkw={x.replace(' ', '+')}&_sacat=0&rt=nc&LH_Sold=1&LH_Complete=1"
            )
            create_pdf_report(excel_path, pdf_path, df)
            
            
            
            print("PDF conversion complete")

            try:
                print("Sending files to user")
                # Send CSV file
                if csv_file_path:
                    await user.send(file=discord.File(csv_file_path))
                
                # Send Excel file
                if excel_path:
                    await user.send(file=discord.File(excel_path))
                
                # Send PDF file
                if pdf_path:
                    await user.send(file=discord.File(pdf_path))
                
                # Send confirmation message
                await user.send("You will be sent your report every day at 12am EST")
                print("Files sent successfully")
                
                # Increment notification counters
                global successful_notifications
                successful_notifications += 1
                total_notifications += 1
                print("Cleaning up files")
                delete_file(csv_file_path)
                delete_file(excel_path)
                delete_file(pdf_path)
                print("Cleanup complete")
            except FileNotFoundError as e:
                print(f"File not found: {e}")
                await user.send("Error: One or more files could not be found.")
            except discord.HTTPException as e:
                print(f"Discord HTTP error: {e}")
                await user.send("Error: Failed to send the files due to a Discord issue.")
            except Exception as e:
                print(f"Unexpected error: {e}")
                await user.send("An unexpected error occurred while sending your files.")

           
            
            
           
            
        except Exception as e:
            print(f"Error in file conversion/sending: {e}")
            raise
            
    except Exception as e:
        print(f"Error in add_percentage_and_send: {e}")
        raise
        
async def delete_previous_bot_messages(channel):
    async for message in channel.history(limit=100):
        if message.author == bot.user:
            await message.delete()
            
            
allowed_roles_rebuild = os.getenv('ALLOWED_ROLES_REBUILD').split(',')
allowed_roles_setzip = os.getenv('ALLOWED_ROLES_SETZIP').split(',')
allowed_roles_setzipbutton = os.getenv('ALLOWED_ROLES_SETZIPBUTTON').split(',')
allowed_roles_setstoreidbutton = os.getenv('ALLOWED_ROLES_SETSTOREIDBUTTON').split(',')

async def send_zip_code_prompt():
    channel = bot.get_channel(zip_code_channel_id)
    if channel:
        await delete_previous_bot_messages(channel)
        embed = discord.Embed(
            title="Set Your Zip Code",
            description="Click the button below to set your zip code \n\n__**You will automatically get sent ALL Walmart leads everyday at 12am EST within 50 miles of you that fulfill this criteria:**__\n- 50% off\n- In stock\n- Posted Within 24 hours\n\n*This is so fire if you are ever going to get groceries, can easily secure a few easy cooks as well at your store (if it says interaction faield, wait a moment then retry the button)*",

            color=discord.Color.blue()
        )
        await channel.send(embed=embed, view=ZipCodeButton())

async def send_store_id_prompt():
    channel = bot.get_channel(store_id_channel_id)
    if channel:
        await delete_previous_bot_messages(channel)
        embed = discord.Embed(
            title="Set Your Store ID",
            description="Click the button below to set your store ID and receive deals in your area.",
            color=discord.Color.green()
        )
        await channel.send(embed=embed, view=StoreIDButton())

class RetryNotificationButton(ui.View):
    def __init__(self, failed_users, timeout=86399):  
        super().__init__(timeout=timeout)
        self.failed_users = failed_users

    @ui.button(label="Retry Failed Notifications", style=discord.ButtonStyle.primary)
    async def retry_notifications(self, interaction: discord.Interaction, button: ui.Button):
        await interaction.response.defer(ephemeral=True)
        
        retry_results = []
        for username in self.failed_users:
            try:
                user = await fetch_user_by_name(username)
                if user:
                    csv_path = os.path.join(user_csv_directory, f'user_{username}_deals.csv')
                    excel_path = csv_path.replace('.csv', '.xlsx')
                    pdf_path = csv_path.replace('.csv', '.pdf')
                    # Test DM permissions first
                    try:
                        await user.send(file=discord.File(csv_path))
                        await user.send(file=discord.File(excel_path))
                        await user.send(file=discord.File(pdf_path))
                        await user.send("You will be sent your report every day at 9am est")
                        retry_results.append(f"✅ {username}: Successfully resent and verified delivery")
                    except discord.Forbidden:
                        retry_results.append(f"❌ {username}: Cannot send - DMs are disabled")
                    except Exception as e:
                        retry_results.append(f"❌ {username}: Delivery failed - {str(e)}")
                else:
                    retry_results.append(f"❌ {username}: User not found")
            except Exception as e:
                retry_results.append(f"❌ {username}: {str(e)}")
        
        results_message = "\n".join(retry_results)
        await interaction.followup.send(f"Retry Results:\n{results_message}", ephemeral=True)

class ZipCodeButton(ui.View):
    def __init__(self, timeout=86400):  # Set timeout to 24 hours
        super().__init__(timeout=timeout)

    @ui.button(label="Set Zip Code", style=discord.ButtonStyle.primary)
    async def set_zip_code(self, interaction: Interaction, button: ui.Button):
        try:
            await interaction.response.send_modal(ZipCodeModal())
        except discord.errors.NotFound:
            try:
                await interaction.followup.send("The interaction expired. Please try clicking the button again.", ephemeral=True)
            except:
                pass
        except Exception as e:
            print(f"Error in set_zip_code button: {e}")
            try:
                await interaction.followup.send("An error occurred. Please try again.", ephemeral=True)
            except:
                pass

class StoreIDButton(ui.View):
    def __init__(self, timeout=86400):  # Set timeout to 24 hours
        super().__init__(timeout=timeout)

    @ui.button(label="Set Store ID", style=discord.ButtonStyle.primary)
    async def set_store_id(self, interaction: Interaction, button: ui.Button):
        try:
            await interaction.response.send_modal(StoreIDModal())
        except discord.errors.NotFound:
            try:
                await interaction.followup.send("The interaction expired. Please try clicking the button again.", ephemeral=True)
            except:
                pass
        except Exception as e:
            print(f"Error in set_store_id button: {e}")
            try:
                await interaction.followup.send("An error occurred. Please try again.", ephemeral=True)
            except:
                pass

class ZipCodeModal(ui.Modal, title="Enter Your Zip Code"):
    zip_code = ui.TextInput(label="Zip Code", placeholder="Enter your zip code here", min_length=5, max_length=5)

    async def on_submit(self, interaction: Interaction):
        try:
            user_roles = [str(role.id) for role in interaction.user.roles]
            

            await interaction.response.defer(ephemeral=True)
            username = interaction.user.name
            zip_code = self.zip_code.value

            if not zip_code.isdigit() or len(zip_code) != 5:
                await interaction.followup.send("Please enter a valid 5-digit zip code.", ephemeral=True)
                return

            conn = sqlite3.connect('user_data.db')
            cursor = conn.cursor()

            try:
                cursor.execute('''
                    INSERT INTO users (username, zip_code) VALUES (?, ?)
                    ON CONFLICT(username) DO UPDATE SET zip_code=excluded.zip_code
                ''', (username, zip_code))
                conn.commit()
            except Exception as e:
                print(f"Database error: {e}")
                await interaction.followup.send("An error occurred while saving your zip code. Please try again.", ephemeral=True)
                return
            finally:
                conn.close()

            # Successfully saved ZIP code
            await interaction.followup.send(f"✅ Zip code {zip_code} saved successfully!\n\nYou will receive daily deal notifications at 12am EST with Walmart deals within 50 miles of your location.", ephemeral=True)

        except Exception as e:
            print(f"Error in zip code modal: {e}")
            try:
                await interaction.followup.send("An error occurred. Please try again.", ephemeral=True)
            except:
                pass

class StoreIDModal(ui.Modal, title="Enter Your Store ID"):
    store_id = ui.TextInput(label="Store ID", placeholder="Enter your store ID here (not ZIP code)", min_length=1, max_length=4)

    async def on_submit(self, interaction: Interaction):
        try:
            user_roles = [str(role.id) for role in interaction.user.roles]
            if not any(role_id in allowed_roles_setstoreidbutton for role_id in user_roles):
                await interaction.response.send_message("You do not have the required role to set the store ID.", ephemeral=True)
                return

            await interaction.response.defer(ephemeral=True)
            username = interaction.user.name
            store_id = self.store_id.value

            if not store_id.isdigit() or len(store_id) > 4:
                await interaction.followup.send("Please enter a valid store ID (up to 4 digits).", ephemeral=True)
                return

            conn = sqlite3.connect('user_data.db')
            cursor = conn.cursor()

            try:
                cursor.execute('''
                    SELECT last_set_time FROM store_users WHERE username = ?
                ''', (username,))
                result = cursor.fetchone()
                
                if result:
                    last_set_time = datetime.strptime(result[0], '%Y-%m-%d %H:%M:%S')
                    if datetime.now() - last_set_time < timedelta(hours=6):
                        await interaction.followup.send("You can only set your store ID once every 6 hours.", ephemeral=True)
                        return

                cursor.execute('''
                    INSERT INTO store_users (username, store_id, last_set_time) 
                    VALUES (?, ?, ?)
                    ON CONFLICT(username) 
                    DO UPDATE SET 
                        store_id=excluded.store_id, 
                        last_set_time=excluded.last_set_time
                ''', (username, store_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
                conn.commit()

            except Exception as e:
                print(f"Database error: {e}")
                await interaction.followup.send("An error occurred while saving your store ID. Please try again.", ephemeral=True)
                return
            finally:
                conn.close()

            # Successfully saved Store ID
            await interaction.followup.send(f"✅ Store ID {store_id} saved successfully!\n\nYou will receive daily deal notifications at 12am EST with Walmart deals from your specific store.", ephemeral=True)

        except Exception as e:
            print(f"Error in store ID modal: {e}")
            try:
                await interaction.followup.send("An error occurred. Please try again.", ephemeral=True)
            except:
                pass

# Command to set the user's zip code and find deals
# Get allowed roles from the .env file
allowed_roles_rebuild = os.getenv('ALLOWED_ROLES_REBUILD').split(',')
allowed_roles_setzip = os.getenv('ALLOWED_ROLES_SETZIP').split(',')
allowed_roles_setzipbutton = os.getenv('ALLOWED_ROLES_SETZIPBUTTON').split(',')
allowed_roles_setstoreidbutton = os.getenv('ALLOWED_ROLES_SETSTOREIDBUTTON').split(',')
allowed_roles_addbuser = os.getenv('ALLOWED_ROLES_ADDDBUSER').split(',')

# Command to add a user to the database manually
@bot.tree.command(name="adddbuser", description="Add a user to the database manually")
async def adddbuser(interaction: discord.Interaction, username: str, zip_code: str):
    # Check if the user has the required role
    user_roles = [str(role.id) for role in interaction.user.roles]
    if not any(role_id in allowed_roles_addbuser for role_id in user_roles):
        await interaction.response.send_message("You do not have the required role to use this command.", ephemeral=True)
        return
    
    # Add or update the user in the database
    conn = sqlite3.connect('user_data.db')
    cursor = conn.cursor()
    cursor.execute('''INSERT INTO users (username, zip_code) VALUES (?, ?)
                      ON CONFLICT(username) DO UPDATE SET zip_code=excluded.zip_code''', (username, zip_code))
    conn.commit()
    conn.close()
    await interaction.response.send_message(f"User {username} with zip code {zip_code} added to the database.", ephemeral=True)

@bot.tree.command(name="dbclear", description="Clear a specific database table")
async def dbclear(interaction: discord.Interaction, table_name: str):
    user_roles = [str(role.id) for role in interaction.user.roles]
    if not any(role_id in allowed_roles_addbuser for role_id in user_roles):
        await interaction.response.send_message("You do not have the required role to use this command.", ephemeral=True)
        return
        
    conn = sqlite3.connect('user_data.db')
    cursor = conn.cursor()
    
    # Get list of valid tables
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
    valid_tables = [table[0] for table in cursor.fetchall()]
    
    if table_name not in valid_tables:
        await interaction.response.send_message(f"Invalid table name. Valid tables are: {', '.join(valid_tables)}", ephemeral=True)
        return
        
    cursor.execute(f"DELETE FROM {table_name}")
    conn.commit()
    conn.close()
    
    await interaction.response.send_message(f"Table '{table_name}' has been cleared.", ephemeral=True)

@bot.tree.command(name="usercount", description="Get count of users in database tables")
async def usercount(interaction: discord.Interaction):
    user_roles = [str(role.id) for role in interaction.user.roles]
    if not any(role_id in allowed_roles_addbuser for role_id in user_roles):
        await interaction.response.send_message("You do not have the required role to use this command.", ephemeral=True)
        return
        
    conn = sqlite3.connect('user_data.db')
    cursor = conn.cursor()
    
    # Get counts from each table
    cursor.execute("SELECT COUNT(*) FROM users")
    zip_users = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM store_users")
    store_users = cursor.fetchone()[0]
    
    conn.close()
    
    embed = discord.Embed(
        title="Bot Usage Statistics",
        color=discord.Color.blue()
    )
    embed.add_field(name="Users with ZIP Codes", value=str(zip_users), inline=True)
    embed.add_field(name="Users with Store IDs", value=str(store_users), inline=True)
    embed.add_field(name="Total Users", value=str(zip_users + store_users), inline=False)
    
    await interaction.response.send_message(embed=embed, ephemeral=True)
    
    
# Command to check deals for a specific store ID
@bot.tree.command(name="storeid", description="Check deals for a specific store ID")
async def storeid(interaction: discord.Interaction, store_id: str):
     # Check if the user has set the store ID within the last 6 hours
    username = interaction.user.name
    # Initialize the database and create tables if they don't exist
    conn = sqlite3.connect('user_data.db')
    cursor = conn.cursor()
    cursor.execute('''
        SELECT last_set_time FROM store_users WHERE username = ?
    ''', (username,))
    result = cursor.fetchone()
    if result:
        last_set_time = datetime.strptime(result[0], '%Y-%m-%d %H:%M:%S')
        if datetime.now() - last_set_time < timedelta(hours=6):
            await interaction.response.send_message("You can only set your store ID once every 6 hours.", ephemeral=True)
            conn.close()
            return

    # Insert or update the user's store ID and update the last set time
    cursor.execute('''
        INSERT INTO store_users (username, store_id, last_set_time) VALUES (?, ?, ?)
        ON CONFLICT(username) DO UPDATE SET store_id=excluded.store_id, last_set_time=excluded.last_set_time
    ''', (username, store_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    conn.commit()
    conn.close()
    # Load the combined CSV file
    try:
        combined_df = pd.read_csv(combined_csv_path, dtype={'Store ID': str})
    except Exception as e:
        await interaction.response.send_message(f"Error loading CSV file: {e}", ephemeral=True)
        return
    
    # Filter the combined CSV for deals in the specified store ID
    
    # Split discount values and convert them back to int, handling 'None' and other non-numeric values
    def convert_discount(x):
        try:
            return int(x.rstrip('%'))
        except (ValueError, AttributeError):
            return 0

    combined_df['discount'] = combined_df['discount'].apply(convert_discount)
    # Convert stock columns to numeric, assuming non-numeric values are 0
    stock_columns = [' Floor Stock', ' Backroom Stock', ' In Transit Stock']
    for col in stock_columns:
        combined_df[col] = pd.to_numeric(combined_df[col], errors='coerce').fillna(0).astype(int)
    filtered_df = combined_df[
    (combined_df['Store ID'] == store_id) &
    (combined_df['discount'] >= 50) &
    (
        (combined_df[' Floor Stock'] > 1) |
        (combined_df[' Backroom Stock'] > 1) |
        (combined_df[' In Transit Stock'] > 1)
    )
    ]
    
    if not filtered_df.empty:
        # Save the filtered DataFrame to a new CSV file in the user_csv_directory
        output_file_path = os.path.join(user_csv_directory, f'store_{store_id}_deals.csv')
        filtered_df.to_csv(output_file_path, index=False)
        
        # Add '%' to the discount column and send the CSV file
        await add_percentage_and_send(output_file_path, interaction.user)
        
        
    else:
        await interaction.response.send_message(f"No deals found for store ID {store_id}.", ephemeral=True)


# Custom check to see if the user has any of the allowed roles
def has_allowed_role(allowed_roles):
    async def predicate(interaction: discord.Interaction):
        user_roles = [role.name for role in interaction.user.roles]
        return any(role in allowed_roles for role in user_roles)
    return commands.check(predicate)

@bot.tree.command(name="rebuild", description="Rebuild the database and combine CSV files")
async def rebuild(interaction: discord.Interaction):
    combined_csv_path_temp = os.path.join(downloads_directory, 'combined_temp.csv')
    channel = bot.get_channel(deals_channel_id)
    clean_downloads_directory()
    async for message in channel.history(limit=100):
        if message.attachments and any(att.filename.endswith('.csv') for att in message.attachments):
            await interaction.response.send_message("Database is being rebuilt. Confirmation will be sent in a DM.", ephemeral=True)
            await fetch_csv_files_from_history(channel)
            process_and_combine_csv(downloads_directory, combined_csv_path_temp)
            reorganize_columns(combined_csv_path_temp, combined_csv_path_temp2)
            update_store_ids(combined_csv_path_temp2, combined_csv_path)
            user = await bot.fetch_user(interaction.user.id)
            await user.send(f"Database has been Rebuilt")
                

# Error handler for missing role
@rebuild.error
async def rebuild_error(interaction: discord.Interaction, error):
    if isinstance(error, commands.CheckFailure):
        await interaction.response.send_message("You do not have the required role to use this command.", ephemeral=True)

@bot.tree.command(name="setzip", description="Set your zip code")
@has_allowed_role(allowed_roles_setzip)
async def setzip(interaction: discord.Interaction, zip_code: str):
    username = interaction.user.name

    # Insert or update the user's zip code in the database
    conn = sqlite3.connect('user_data.db')
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO users (username, zip_code) VALUES (?, ?)
        ON CONFLICT(username) DO UPDATE SET zip_code=excluded.zip_code
    ''', (username, zip_code))
    conn.commit()
    conn.close()

    user = await bot.fetch_user(interaction.user.id)
    await user.send(f"Zip code {zip_code} saved for user {username}.")

    await interaction.response.send_message(f"Zip code {zip_code} saved for user {username}.\nYou will be sent your report every sunday at 12am est", ephemeral=True)

    # Load the combined CSV file with ZIP codes as strings to avoid float conversion issues
    print(f"Got past zip code")
    try:
        combined_df = pd.read_csv(combined_csv_path)
        print("CSV file loaded successfully.")
        print(combined_df.head())
        print("Column names in the CSV file:")
        print(combined_df.columns)
    except Exception as e:
        print(f"Error loading CSV file: {e}")
        return

    # Debugging: Show the unique ZIP codes in combined.csv as strings
    unique_zip_codes = combined_df[' ZIP'].unique()
    for zip_codeunique in unique_zip_codes:
        print(zip_codeunique)

    # Ensure ZIP codes are compared as strings, remove any leading/trailing spaces, and strip the period and anything after
    combined_df[' ZIP'] = combined_df[' ZIP'].astype(str).str.strip().str.split('.').str[0]

    # Use the storecheck function to find nearby ZIP codes
    radius = 50  # miles
    nearby_zip_codes = storecheck(zip_code, radius)

    # Debugging: Show the ZIP codes within the radius as strings
    print(f"ZIP codes within {radius} miles of {zip_code}: {nearby_zip_codes}")

    nearby_zip_codes = [str(z).strip().split('.')[0] for z in nearby_zip_codes]

    # Filter the combined CSV for deals in the nearby ZIP codes, comparing only the first 5 numbers
    # Split discount values and convert them back to int, handling 'None' and other non-numeric values
    def convert_discount(x):
        try:
            return int(x.rstrip('%'))
        except (ValueError, AttributeError):
            return 0

    combined_df['discount'] = combined_df['discount'].apply(convert_discount)
    # Convert stock columns to numeric, assuming non-numeric values are 0
    stock_columns = [' Floor Stock', ' Backroom Stock', ' In Transit Stock']
    for col in stock_columns:
        combined_df[col] = pd.to_numeric(combined_df[col], errors='coerce').fillna(0).astype(int)
    filtered_df = combined_df[
    (combined_df[' ZIP'].str[:5].isin([z[:5] for z in nearby_zip_codes])) &
    (combined_df['discount'] >= 50) &
    (
        (combined_df[' Floor Stock'] > 1) |
        (combined_df[' Backroom Stock'] > 1) |
        (combined_df[' In Transit Stock'] > 1)
    )
    ]
    
    if not filtered_df.empty:
        # Save the filtered DataFrame to a new CSV file in the user_csv_directory
        output_file_path = os.path.join(user_csv_directory, f'user_{username}_deals.csv')
        filtered_df.to_csv(output_file_path, index=False)
        
        # Add '%' to the discount column and send the CSV file
        await add_percentage_and_send(output_file_path, user)
        
        
    else:
        no_deals_message = "No deals found within your radius."
        await user.send(no_deals_message)
        
        
zip_code_channel_id = int(os.getenv('ZIP_CODE_CHANNEL_ID'))  # Replace with your zip code channel ID
store_id_channel_id = int(os.getenv('STORE_ID_CHANNEL_ID'))  # Replace with your store ID channel ID

@tasks.loop(hours=23)
async def send_prompts():
    await send_zip_code_prompt()
    await send_store_id_prompt()
@tasks.loop(minutes=1)
async def daily_deals_notification():
    eastern = pytz.timezone('US/Eastern')
    now = datetime.now(eastern)
    
    print(f"Current time: {now}")
    # Check if it's 9 AM
    start_time = now.replace(hour=18, minute=0, second=0, microsecond=0)
    end_time = now.replace(hour=18, minute=1, second=0, microsecond=0)

        # Calculate and print time until start_time
    if now < start_time:
        time_until = start_time - now
        print(f"Time until next notification: {time_until}")
    
    if start_time <= now <= end_time:
        print(f"Daily Notification Running")
        await notify_users_of_new_deals()


@tasks.loop(hours=24)
async def auto_rebuild():
    from datetime import datetime, timedelta
    print("Starting auto-rebuild...")
    clean_downloads_directory()
    # 2. Proceed with rebuild
    combined_csv_path_temp = os.path.join(downloads_directory, 'combined_temp.csv')
    channel = bot.get_channel(deals_channel_id)
    async for message in channel.history(limit=100):
        if message.attachments and any(att.filename.endswith('.csv') for att in message.attachments):
            await fetch_csv_files_from_history(channel)
            process_and_combine_csv(downloads_directory, combined_csv_path_temp)
            reorganize_columns(combined_csv_path_temp, combined_csv_path_temp2)
            update_store_ids(combined_csv_path_temp2, combined_csv_path)
            print("Auto-rebuild complete")
            break
    

# Run the bot
bot.run(os.getenv('DISCORD_BOT_TOKEN'))